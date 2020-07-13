import argparse
from openpyxl import load_workbook
import geocoder

# Construct the argument parser
ap = argparse.ArgumentParser()

# Add the arguments to the parser
ap.add_argument("-i", type=str, dest='input', required=True, help="input file with addresses to process")
ap.add_argument("-w", type=str, dest='worksheet', required=True, help="worksheet to process")
ap.add_argument("-o", type=str, dest='output', required=True, help="output file where input file and GPS coordination will be stored")
ap.add_argument("-k", type=str, dest='key', required=True, help="GCP API key")
args = ap.parse_args()

API_KEY = args.key
INPUT_FILE = args.input
OUTPUT_FILE = args.output
WORKSHEET = args.worksheet

# helper method to get cell value as string
def get_cell(sheet, col_id, row_id):
    return str(sheet[f'{col_id}{row_id}'].value)

# helper method to get POI address as a string
def get_adr(sheet, r):
    row_id = str(r)
    adr = get_cell(sheet, 'H', row_id) + ' ' + get_cell(sheet, 'I', row_id) + ', ' + get_cell(sheet, 'K', row_id) + ' ' + get_cell(sheet, 'L', row_id) + ', ' + get_cell(sheet, 'M', row_id)
    return adr

# Open workbook to process
wb = load_workbook(filename=INPUT_FILE)
ws = wb[WORKSHEET]
# Count max number of rows
max_row_for_a = max((a.row for a in ws['A'] if a.value is not None))

print(f'Reading file {INPUT_FILE}... total number of rows to procced is {max_row_for_a}')

for r in range(2, max_row_for_a+1):
    adr = get_adr(ws,r)
    g = geocoder.google(adr, key=API_KEY) # Ask google API for GPS coordinates of the address
    ws['B'+str(r)] = g.latlng[0]
    ws['C'+str(r)] = g.latlng[1]
    print(f'{g.latlng}: {adr}')

# Append GPS coordinates to workbook and save as new file
wb.save(OUTPUT_FILE)
print(f'Output file is {OUTPUT_FILE}. Geocoding done.')
